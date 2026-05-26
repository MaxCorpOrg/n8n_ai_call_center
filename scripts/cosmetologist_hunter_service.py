#!/usr/bin/env python3
import argparse
from collections import deque
import json
import os
import re
import shutil
import subprocess
import threading
import time
import urllib.parse
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from socketserver import ThreadingMixIn

import requests
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent


def load_env_file_if_exists(path):
    env_path = Path(path)
    if not env_path.exists():
        return
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


for candidate in (PROJECT_ROOT / ".env.cosmetologist_hunter", Path.cwd() / ".env.cosmetologist_hunter"):
    load_env_file_if_exists(candidate)

DEFAULT_TABLES_DIR = PROJECT_ROOT / " Таблицы_контактов "
DEFAULT_RUNTIME_DIR = PROJECT_ROOT / ".runtime" / "cosmetologist_hunter"

DEFAULT_SOURCE_SPREADSHEET_ID = "1pLrCNeQ_thipr5-fajPusgNZZSd5NHEZFmGkegfpIqI"
DEFAULT_SHEET_NAME = "Лиды_обзвон"
DEFAULT_LOCAL_OUTPUT_DIR = str(DEFAULT_TABLES_DIR)
DEFAULT_TEMPLATE_XLSX = str(DEFAULT_TABLES_DIR / "пример_таблицы.xlsx")
DEFAULT_SETTINGS_PATH = str(DEFAULT_RUNTIME_DIR / "settings.json")
DEFAULT_PREVIEW_DIR = str(DEFAULT_RUNTIME_DIR / "previews")
DEFAULT_DRIVE_FOLDER_ID = os.getenv("COSMETOLOGIST_HUNTER_DRIVE_FOLDER_ID", "").strip()
DEFAULT_SERVER_TOOL_ROOT = os.getenv("COSMETOLOGIST_HUNTER_SERVER_TOOL_ROOT", "/home/aicore/agent-tools").strip()
DEFAULT_FIRECRAWL_ROOT = os.getenv(
    "COSMETOLOGIST_HUNTER_FIRECRAWL_ROOT", str(Path(DEFAULT_SERVER_TOOL_ROOT) / "firecrawl")
).strip()
DEFAULT_FIRECRAWL_BASE_URL = os.getenv("COSMETOLOGIST_HUNTER_FIRECRAWL_BASE_URL", "").strip()
DEFAULT_FIRECRAWL_API_KEY = os.getenv("COSMETOLOGIST_HUNTER_FIRECRAWL_API_KEY", "").strip()
DEFAULT_SITE_CONTROL_ROOT = os.getenv(
    "COSMETOLOGIST_HUNTER_SITE_CONTROL_ROOT", str(Path(DEFAULT_SERVER_TOOL_ROOT) / "site-control-kit")
).strip()
DEFAULT_SITE_CONTROL_SERVER_URL = os.getenv("COSMETOLOGIST_HUNTER_SITE_CONTROL_SERVER_URL", "").strip()
DEFAULT_SITE_CONTROL_TOKEN = os.getenv("COSMETOLOGIST_HUNTER_SITE_CONTROL_TOKEN", "").strip()
DEFAULT_SITE_CONTROL_CLIENT_ID = os.getenv("COSMETOLOGIST_HUNTER_SITE_CONTROL_CLIENT_ID", "").strip()
DEFAULT_SITE_CONTROL_BROWSER_SERVICE = os.getenv(
    "COSMETOLOGIST_HUNTER_SITE_CONTROL_BROWSER_SERVICE", "site-control-kit-browser.service"
).strip()
DEFAULT_SITE_CONTROL_HUB_SERVICE = os.getenv(
    "COSMETOLOGIST_HUNTER_SITE_CONTROL_HUB_SERVICE", "site-control-kit-hub.service"
).strip()
DEFAULT_SITE_CONTROL_BROWSER_TIMER = os.getenv(
    "COSMETOLOGIST_HUNTER_SITE_CONTROL_BROWSER_TIMER", "site-control-kit-browser.timer"
).strip()
DEFAULT_SITE_CONTROL_ON_DEMAND = os.getenv("COSMETOLOGIST_HUNTER_SITE_CONTROL_ON_DEMAND", "1").strip().lower() not in {
    "0",
    "false",
    "no",
    "off",
}
DEFAULT_SEARCH_TIMEOUT_SECONDS = int(os.getenv("COSMETOLOGIST_HUNTER_SEARCH_TIMEOUT_SECONDS", "90") or "90")
DEFAULT_PORT = 8787
MAX_FETCH_TRACE_ITEMS = 120

LEADS_HEADERS = [
    "created_at",
    "updated_at",
    "lead_id",
    "source_system",
    "source_record_key",
    "company_name",
    "contact_name",
    "phone_primary",
    "phone_secondary",
    "city",
    "segment",
    "lpr_role",
    "lpr_confirmed",
    "current_supplier",
    "current_product",
    "current_price",
    "pain_points",
    "objection_code",
    "objection_text",
    "interest_level",
    "call_result",
    "next_step",
    "next_call_at",
    "preferred_channel",
    "manager_owner",
    "expected_volume",
    "expected_budget",
    "material_sent",
    "followup_count",
    "max_touch_limit",
    "do_not_call",
    "final_reason",
    "notes_short",
    "notes_redacted",
    "call_record_url",
    "eleven_conv_id",
    "n8n_execution_id",
    "agent_version",
    "last_updated_by",
]

CITY_PRESETS = {
    "москва": {
        "prodoctorov_city": "moskva",
        "yandex_urls": [
            "https://yandex.com/maps/213/moscow/search/{query}/?ll=37.385272%2C55.584227&z=8.9",
        ],
        "d2gis_urls": [
            "https://2gis.ru/moscow/search/{query}/page/{page}?m=37.758667%2C55.681592%2F9.59",
        ],
    },
}

SEARCH_STRATEGIES = {
    "core": [
        "косметолог",
        "косметология",
        "эстетическая косметология",
        "косметологический кабинет",
    ],
    "private": [
        "частный косметолог",
        "частный врач косметолог",
        "косметолог частная практика",
        "косметолог на дому",
        "частный кабинет косметолога",
        "врач косметолог частный кабинет",
        "дерматолог косметолог частная практика",
        "косметолог инъекции частный",
    ],
    "doctor": [
        "врач косметолог",
        "врач-косметолог",
        "дерматолог косметолог",
        "врач дерматолог косметолог",
        "инъекционная косметология",
        "косметолог эстетист",
        "дерматовенеролог косметолог",
        "трихолог косметолог",
    ],
    "clinic": [
        "клиника косметологии",
        "центр косметологии",
        "клиника эстетической медицины",
        "центр эстетической медицины",
        "косметологическая клиника",
    ],
}

STRICT_NAME_KEYS = [
    "космет",
    "эстет",
    "дермат",
    "медиц",
    "clinic",
    "doctor",
    "лазер",
    "laser",
    "эпил",
    "beauty",
    "бьюти",
    "трихолог",
    "перманент",
]

STRICT_RUBRIC_ALIASES = {
    "kosmetolog",
    "ehpilyaciya",
    "permanentnyjj_makiyazh",
}

PRIVATE_POSITIVE_KEYS = [
    "частн",
    "частная практика",
    "на дому",
    "кабинет косметолога",
    "врач косметолог",
    "врач-косметолог",
    "дерматолог косметолог",
    "дерматолог-косметолог",
    "косметолог эстетист",
    "ип ",
]

PRIVATE_ORG_REJECT_KEYS = [
    "клиник",
    "медицинский центр",
    "медцентр",
    "центр косметологии",
    "центр эстет",
    "центр красоты",
    "салон",
    "студия",
    "spa",
    "спа",
    "beauty",
    "бьюти",
    "пространство красоты",
    "многопроф",
    "стоматолог",
    "лаборатор",
    "ооо",
    "зао",
    "ао ",
]
GOOGLE_AUTH_RETRY_MARKERS = (
    "Method doesn't allow unregistered callers",
    "PERMISSION_DENIED",
)


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_parent(path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def sanitize_title_part(value):
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^0-9a-zа-я_]+", "", text)
    return text or "gorod"


def build_title_prefix(city):
    return f"контакты_косметологов_{sanitize_title_part(city)}"


def build_sheet_title(city, ordinal):
    return f"{build_title_prefix(city)}_{ordinal}"


def normalize_phone(value):
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) == 10:
        digits = "7" + digits
    elif len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    elif len(digits) > 11 and digits.startswith("7"):
        digits = digits[:11]
    if len(digits) == 11 and digits.startswith("7"):
        return f"+{digits}"
    return ""


def normalize_company_name(value):
    text = fix_text(value or "").lower().replace("ё", "е")
    text = re.sub(r"[\"'`«»]", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_address(value):
    text = fix_text(value or "").lower().replace("ё", "е")
    text = re.sub(r"[\"'`«»]", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def make_company_address_key(company_name, address):
    company_key = normalize_company_name(company_name)
    address_key = normalize_address(address)
    if company_key and address_key:
        return f"{company_key} | {address_key}"
    return ""


def extract_stored_address(notes_short):
    text = fix_text(notes_short or "").strip()
    match = re.match(r"^Адрес:\s*(.+)$", text, re.I)
    return match.group(1).strip() if match else ""


def unique_phones(values):
    phones = []
    seen = set()
    for value in values or []:
        phone = normalize_phone(value)
        if phone and phone not in seen:
            seen.add(phone)
            phones.append(phone)
    return phones


def extract_json_object_after(html, needle):
    text = str(html or "")
    idx = text.find(needle)
    if idx == -1:
        return {}
    payload = text[idx + len(needle) :]
    brace = 0
    in_string = False
    escaped = False
    end = None
    for index, char in enumerate(payload):
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
        else:
            if char == '"':
                in_string = True
            elif char == "{":
                brace += 1
            elif char == "}":
                brace -= 1
                if brace == 0:
                    end = index + 1
                    break
    if not end:
        return {}
    try:
        return json.loads(payload[:end])
    except Exception:
        return {}


def walk_json_like(root):
    stack = [root]
    while stack:
        current = stack.pop()
        yield current
        if isinstance(current, dict):
            stack.extend(current.values())
        elif isinstance(current, list):
            stack.extend(current)


def fix_text(value):
    if not isinstance(value, str):
        return value
    value = value.strip()
    if not value:
        return value
    try:
        fixed = value.encode("latin1").decode("utf-8")
        if fixed.count("\ufffd") <= value.count("\ufffd"):
            return fixed
    except Exception:
        pass
    return value


def relevant_text(value):
    text = fix_text(value or "").lower()
    return any(key in text for key in STRICT_NAME_KEYS)


def looks_like_person_name(value):
    text = fix_text(value or "").strip()
    parts = re.findall(r"\b[А-ЯЁ][а-яё]{2,}\b", text)
    return len(parts) >= 2


def private_cosmetologist_match(item):
    name = fix_text(item.get("company_name") or "").strip()
    source = str(item.get("source") or "").strip().lower()
    haystack = " ".join(
        fix_text(value or "").strip()
        for value in [
            name,
            item.get("address") or "",
            item.get("private_hint") or "",
            item.get("source_url") or "",
        ]
    ).lower()

    if not ("космет" in haystack or "дермат" in haystack or source == "prodoctorov"):
        return False, "no_cosmetology_signal"

    if source == "prodoctorov" and looks_like_person_name(name):
        return True, "doctor_profile"

    if any(key in haystack for key in PRIVATE_ORG_REJECT_KEYS):
        if "кабинет косметолога" not in haystack and "частн" not in haystack and not looks_like_person_name(name):
            return False, "organization_keyword"

    if any(key in haystack for key in PRIVATE_POSITIVE_KEYS):
        return True, "private_keyword"

    if looks_like_person_name(name):
        return True, "person_name"

    return False, "not_private_enough"


def score_name(name):
    text = str(name or "").lower()
    score = 0
    if "космет" in text:
        score += 60
    if "дермат" in text:
        score += 25
    if "эстет" in text:
        score += 40
    if "медиц" in text or "clinic" in text:
        score += 20
    if "лазер" in text or "laser" in text:
        score += 25
    if "эпил" in text:
        score += 20
    if "трихолог" in text:
        score += 15
    if "beauty" in text or "бьюти" in text:
        score += 15
    if "салон красоты" in text:
        score -= 10
    if "семейн" in text or "многопроф" in text or "добромед" in text:
        score -= 30
    if "пространство красоты" in text:
        score -= 10
    return score


def build_output_rows(records):
    timestamp = now_text()
    rows = []
    for row_idx, rec in enumerate(records, start=2):
        row = {header: "" for header in LEADS_HEADERS}
        address = fix_text(rec.get("address") or "").strip()
        source = str(rec.get("source") or "").strip()
        source_url = str(rec.get("source_url") or "").strip()
        notes_redacted = ""
        if source or source_url:
            parts = []
            if source:
                parts.append(f"Источник: {source}")
            if source_url:
                parts.append(f"URL: {source_url}")
            notes_redacted = " | ".join(parts)
        row.update(
            {
                "created_at": timestamp,
                "updated_at": timestamp,
                "lead_id": "",
                "source_system": "xlsx_import",
                "source_record_key": f"row_{row_idx}",
                "company_name": rec.get("company_name") or "Косметолог",
                "contact_name": "",
                "phone_primary": rec["phone_primary"],
                "phone_secondary": rec.get("phone_secondary") or "",
                "city": rec.get("city") or "",
                "segment": "Косметолог",
                "followup_count": "0",
                "max_touch_limit": "3",
                "do_not_call": "false",
                "notes_short": f"Адрес: {address}" if address else "",
                "notes_redacted": notes_redacted,
                "agent_version": "AI_CALL_AGENT_1",
                "last_updated_by": "system_seed",
            }
        )
        rows.append([row.get(header, "") for header in LEADS_HEADERS])
    return rows


def load_json_file(path, default):
    file_path = Path(path)
    if not file_path.exists():
        return default
    return json.loads(file_path.read_text(encoding="utf-8"))


def save_json_file(path, payload):
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def default_google_source_candidates():
    return [
        os.getenv("GOOGLE_OAUTH_SOURCE_JSON", ""),
        str(PROJECT_ROOT / "backups" / "2026-04-07_human_gate_autodial_refresh" / "autodial_live_after.json"),
        str(PROJECT_ROOT / "backups" / "2026-04-06_live_autodial_gid_sync" / "call_log_before_20260406_154135.json"),
    ]


def extract_google_oauth_from_text(text):
    client_id = re.search(r"client_id:\s*'([^']+)'", text)
    client_secret = re.search(r"client_secret:\s*'([^']+)'", text)
    refresh_token = re.search(r"refresh_token:\s*'([^']+)'", text)
    if not (client_id and client_secret and refresh_token):
        return None
    return {
        "client_id": client_id.group(1),
        "client_secret": client_secret.group(1),
        "refresh_token": refresh_token.group(1),
    }


def load_google_oauth():
    direct = {
        "client_id": os.getenv("GOOGLE_CLIENT_ID", "").strip(),
        "client_secret": os.getenv("GOOGLE_CLIENT_SECRET", "").strip(),
        "refresh_token": os.getenv("GOOGLE_REFRESH_TOKEN", "").strip(),
    }
    if all(direct.values()):
        return direct

    for candidate in default_google_source_candidates():
        if not candidate:
            continue
        path = Path(candidate)
        if not path.exists():
            continue
        creds = extract_google_oauth_from_text(path.read_text(encoding="utf-8", errors="ignore"))
        if creds:
            return creds

    raise RuntimeError(
        "Не найдены Google OAuth credentials. Задайте GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET / GOOGLE_REFRESH_TOKEN."
    )


class FirecrawlClient:
    def __init__(self, base_url="", api_key="", root_path=""):
        self.base_url = str(base_url or "").rstrip("/")
        self.api_key = str(api_key or "").strip()
        self.root_path = str(root_path or "").strip()
        self.session = requests.Session()

    def enabled(self):
        return bool(self.base_url)

    def scrape_html(self, url, timeout=90):
        if not self.enabled():
            return ""
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        payload = {
            "url": url,
            "formats": ["rawHtml", "html"],
            "onlyMainContent": False,
            "waitFor": 1500,
            "timeout": int(timeout * 1000),
        }
        try:
            resp = self.session.post(f"{self.base_url}/v2/scrape", headers=headers, json=payload, timeout=timeout)
            resp.raise_for_status()
            body = resp.json()
        except Exception:
            return ""
        data = body.get("data") or {}
        return str(data.get("rawHtml") or data.get("html") or "")

    def status(self):
        return {
            "enabled": self.enabled(),
            "base_url": self.base_url,
            "root_path": self.root_path,
            "has_api_key": bool(self.api_key),
        }


class SiteControlKitClient:
    def __init__(
        self,
        server_url="",
        token="",
        client_id="",
        root_path="",
        browser_service="",
        hub_service="",
        browser_timer="",
        on_demand=True,
    ):
        self.server_url = str(server_url or "").rstrip("/")
        self.token = str(token or "").strip()
        self.client_id = str(client_id or "").strip()
        self.root_path = str(root_path or "").strip()
        self.browser_service = str(browser_service or DEFAULT_SITE_CONTROL_BROWSER_SERVICE).strip()
        self.hub_service = str(hub_service or DEFAULT_SITE_CONTROL_HUB_SERVICE).strip()
        self.browser_timer = str(browser_timer or DEFAULT_SITE_CONTROL_BROWSER_TIMER).strip()
        self.on_demand = bool(on_demand)
        self.session = requests.Session()
        self._clients_cache = []
        self._clients_cache_at = 0.0
        self._started_on_demand = False

    def enabled(self):
        return bool(self.server_url and self.token)

    def _headers(self):
        return {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-Access-Token": self.token,
        }

    def _get(self, path, timeout=30):
        resp = self.session.get(f"{self.server_url}{path}", headers=self._headers(), timeout=timeout)
        resp.raise_for_status()
        return resp.json()

    def _post(self, path, payload, timeout=30):
        resp = self.session.post(f"{self.server_url}{path}", headers=self._headers(), json=payload, timeout=timeout)
        resp.raise_for_status()
        return resp.json()

    def _run_systemctl(self, *args, timeout=30):
        if not self.on_demand:
            return False
        if not shutil.which("sudo"):
            return False
        try:
            subprocess.run(
                ["sudo", "-n", "systemctl", *args],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=timeout,
            )
            return True
        except Exception:
            return False

    def _service_state(self, unit_name):
        if not unit_name:
            return "unknown"
        try:
            result = subprocess.run(
                ["systemctl", "is-active", unit_name],
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )
            return (result.stdout or result.stderr or "unknown").strip()
        except Exception:
            return "unknown"

    def ensure_browser_ready(self, timeout=35):
        client = self.pick_client()
        if client:
            return client
        if not self.on_demand:
            return None
        started_hub = self._run_systemctl("start", self.hub_service, timeout=20)
        started_browser = self._run_systemctl("start", self.browser_service, timeout=20)
        if started_hub or started_browser:
            self._started_on_demand = True
        deadline = time.time() + max(5, int(timeout or 35))
        while time.time() < deadline:
            client = self.pick_client()
            if client:
                return client
            time.sleep(1)
        return None

    def shutdown(self):
        if not self.on_demand or not self._started_on_demand:
            return False
        stopped = self._run_systemctl("stop", self.browser_service, timeout=20)
        self._run_systemctl("stop", self.hub_service, timeout=20)
        self._clients_cache = []
        self._clients_cache_at = time.time()
        self._started_on_demand = False
        return stopped

    def list_clients(self, force=False):
        if not self.enabled():
            return []
        if not force and self._clients_cache and (time.time() - self._clients_cache_at) < 5:
            return self._clients_cache
        try:
            body = self._get("/api/clients")
        except Exception:
            self._clients_cache_at = time.time()
            return self._clients_cache
        clients = body.get("clients")
        self._clients_cache = clients if isinstance(clients, list) else []
        self._clients_cache_at = time.time()
        return self._clients_cache

    def pick_client(self):
        clients = self.list_clients()
        if not clients:
            return None
        if self.client_id:
            for client in clients:
                if str(client.get("client_id", "")).strip() == self.client_id:
                    return client
            return None
        return sorted(clients, key=lambda item: str(item.get("last_seen", "")), reverse=True)[0]

    def _wait_command(self, command_id, timeout=45, poll_interval=0.5):
        deadline = time.time() + timeout
        while time.time() < deadline:
            body = self._get(f"/api/commands/{command_id}")
            command = body.get("command") or {}
            if command.get("status") in {"completed", "failed", "cancelled", "timed_out"}:
                return command
            time.sleep(max(poll_interval, 0.1))
        return {}

    def _extract_result(self, command, client_id):
        deliveries = command.get("deliveries") if isinstance(command, dict) else {}
        if not isinstance(deliveries, dict):
            return {}
        delivery = deliveries.get(client_id) or {}
        result = delivery.get("result")
        return result if isinstance(result, dict) else {}

    def send_command(self, target, command, timeout_ms=40000, wait_timeout=45):
        if not self.enabled():
            return {}
        try:
            body = self._post(
                "/api/commands",
                {
                    "issued_by": "cosmetologist-hunter",
                    "timeout_ms": timeout_ms,
                    "target": target,
                    "command": command,
                },
                timeout=max(30, wait_timeout),
            )
        except Exception:
            return {}
        command_id = str(body.get("command_id") or "").strip()
        if not command_id:
            return {}
        return self._wait_command(command_id, timeout=wait_timeout)

    def fetch_html(self, url, wait_timeout=45):
        client = self.ensure_browser_ready(timeout=max(20, wait_timeout))
        if not client:
            return ""
        client_id = str(client.get("client_id", "")).strip()
        if not client_id:
            return ""
        target = {
            "client_id": client_id,
            "active": True,
        }
        navigate = self.send_command(target, {"type": "navigate", "url": url}, timeout_ms=45000, wait_timeout=wait_timeout)
        if not navigate or navigate.get("status") != "completed":
            return ""
        self.send_command(
            target,
            {"type": "wait_selector", "selector": "body", "timeout_ms": 20000, "visible_only": False},
            timeout_ms=25000,
            wait_timeout=wait_timeout,
        )
        html_command = self.send_command(target, {"type": "get_html"}, timeout_ms=25000, wait_timeout=wait_timeout)
        result = self._extract_result(html_command, client_id)
        data = result.get("data") if isinstance(result, dict) else {}
        if not isinstance(data, dict):
            return ""
        return str(data.get("html") or "")

    def status(self):
        clients = self.list_clients()
        selected = self.pick_client()
        return {
            "enabled": self.enabled(),
            "server_url": self.server_url,
            "root_path": self.root_path,
            "client_id": self.client_id,
            "on_demand": self.on_demand,
            "browser_service": self.browser_service,
            "hub_service": self.hub_service,
            "browser_timer": self.browser_timer,
            "browser_service_state": self._service_state(self.browser_service),
            "hub_service_state": self._service_state(self.hub_service),
            "connected_clients": len(clients),
            "selected_client_id": str((selected or {}).get("client_id", "")).strip(),
        }


class GoogleSheetsClient:
    def __init__(self, source_spreadsheet_id, source_sheet_name, drive_folder_id=""):
        self.source_spreadsheet_id = source_spreadsheet_id
        self.source_sheet_name = source_sheet_name
        self.drive_folder_id = str(drive_folder_id or "").strip()
        self.session = requests.Session()
        self._access_token = None

    def access_token(self):
        if self._access_token:
            return self._access_token
        creds = load_google_oauth()
        resp = self.session.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": creds["client_id"],
                "client_secret": creds["client_secret"],
                "refresh_token": creds["refresh_token"],
                "grant_type": "refresh_token",
            },
            timeout=60,
        )
        resp.raise_for_status()
        self._access_token = resp.json()["access_token"]
        return self._access_token

    def refresh_access_token(self):
        self._access_token = None
        return self.access_token()

    def auth_headers(self, json_body=False):
        headers = {"Authorization": f"Bearer {self.access_token()}"}
        if json_body:
            headers["Content-Type"] = "application/json"
        return headers

    def _should_retry_auth(self, response):
        if response is None:
            return False
        text = ""
        try:
            text = response.text or ""
        except Exception:
            text = ""
        return response.status_code in {401, 403} and any(marker in text for marker in GOOGLE_AUTH_RETRY_MARKERS)

    def request_with_auth(self, method, url, *, json_body=False, retry_on_auth=True, **kwargs):
        headers = dict(kwargs.pop("headers", {}) or {})
        headers.update(self.auth_headers(json_body=json_body))
        response = self.session.request(method, url, headers=headers, **kwargs)
        if retry_on_auth and self._should_retry_auth(response):
            self.refresh_access_token()
            headers = dict(kwargs.pop("headers", {}) or {})
            headers.update(self.auth_headers(json_body=json_body))
            response = self.session.request(method, url, headers=headers, **kwargs)
        response.raise_for_status()
        return response

    def read_values(self, spreadsheet_id, sheet_name=None):
        title = sheet_name or self.source_sheet_name
        range_encoded = urllib.parse.quote(f"{title}!A:AM", safe="")
        url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/"
            f"{range_encoded}?majorDimension=ROWS"
        )
        resp = self.request_with_auth("GET", url, timeout=60)
        return resp.json().get("values", [])

    def phones_from_sheet(self, spreadsheet_id, sheet_name=None):
        values = self.read_values(spreadsheet_id, sheet_name)
        if not values:
            return set()
        header = values[0]
        if "phone_primary" not in header:
            return set()
        phone_idx = header.index("phone_primary")
        phones = set()
        for row in values[1:]:
            if phone_idx < len(row):
                phone = normalize_phone(row[phone_idx])
                if phone:
                    phones.add(phone)
        return phones

    def dedup_signatures_from_sheet(self, spreadsheet_id, sheet_name=None):
        values = self.read_values(spreadsheet_id, sheet_name)
        signatures = {
            "phones": set(),
            "names": set(),
            "company_addresses": set(),
        }
        if not values:
            return signatures
        header = values[0]
        index_map = {str(name or "").strip(): idx for idx, name in enumerate(header)}
        phone_primary_idx = index_map.get("phone_primary")
        phone_secondary_idx = index_map.get("phone_secondary")
        company_name_idx = index_map.get("company_name")
        notes_short_idx = index_map.get("notes_short")

        for row in values[1:]:
            if phone_primary_idx is not None and phone_primary_idx < len(row):
                phone = normalize_phone(row[phone_primary_idx])
                if phone:
                    signatures["phones"].add(phone)
            if phone_secondary_idx is not None and phone_secondary_idx < len(row):
                phone = normalize_phone(row[phone_secondary_idx])
                if phone:
                    signatures["phones"].add(phone)
            company_name = ""
            if company_name_idx is not None and company_name_idx < len(row):
                company_name = row[company_name_idx]
                company_key = normalize_company_name(company_name)
                if company_key:
                    signatures["names"].add(company_key)
            notes_short = ""
            if notes_short_idx is not None and notes_short_idx < len(row):
                notes_short = row[notes_short_idx]
            address = extract_stored_address(notes_short)
            company_address_key = make_company_address_key(company_name, address)
            if company_address_key:
                signatures["company_addresses"].add(company_address_key)
        return signatures

    def list_prefixed_sheets(self, title_prefix):
        resp = self.request_with_auth(
            "GET",
            "https://www.googleapis.com/drive/v3/files",
            params={
                "q": (
                    "mimeType='application/vnd.google-apps.spreadsheet' "
                    f"and name contains '{title_prefix}' and trashed=false"
                ),
                "fields": "files(id,name,createdTime)",
                "pageSize": 200,
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
            },
            timeout=60,
        )
        return resp.json().get("files", [])

    def next_ordinal(self, title_prefix):
        ordinals = []
        for item in self.list_prefixed_sheets(title_prefix):
            match = re.search(rf"{re.escape(title_prefix)}_(\d+)$", item.get("name", ""))
            if match:
                ordinals.append(int(match.group(1)))
        return (max(ordinals) + 1) if ordinals else 1

    def existing_signatures_for_city(self, city):
        title_prefix = build_title_prefix(city)
        signatures = self.dedup_signatures_from_sheet(self.source_spreadsheet_id, self.source_sheet_name)
        for item in self.list_prefixed_sheets(title_prefix):
            extra = self.dedup_signatures_from_sheet(item["id"], self.source_sheet_name)
            signatures["phones"].update(extra["phones"])
            signatures["names"].update(extra["names"])
            signatures["company_addresses"].update(extra["company_addresses"])
        return signatures

    def copy_sheet(self, title):
        payload = {"name": title}
        if self.drive_folder_id:
            payload["parents"] = [self.drive_folder_id]
        resp = self.request_with_auth(
            "POST",
            f"https://www.googleapis.com/drive/v3/files/{self.source_spreadsheet_id}/copy",
            params={"supportsAllDrives": "true"},
            json=payload,
            timeout=60,
            json_body=True,
        )
        spreadsheet_id = resp.json()["id"]
        self.rename_sheet_title(spreadsheet_id, title)
        return spreadsheet_id

    def rename_sheet_title(self, spreadsheet_id, title):
        self.request_with_auth(
            "PATCH",
            f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}",
            params={"supportsAllDrives": "true"},
            json={"name": title},
            timeout=60,
            json_body=True,
        )
        self.request_with_auth(
            "POST",
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}:batchUpdate",
            json={
                "requests": [
                    {
                        "updateSpreadsheetProperties": {
                            "properties": {"title": title},
                            "fields": "title",
                        }
                    }
                ]
            },
            timeout=60,
            json_body=True,
        )

    def clear_data_rows(self, spreadsheet_id, sheet_name=None):
        title = sheet_name or self.source_sheet_name
        range_encoded = urllib.parse.quote(f"{title}!A2:AM", safe="")
        self.request_with_auth(
            "POST",
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{range_encoded}:clear",
            json={},
            timeout=60,
            json_body=True,
        )

    def append_rows(self, spreadsheet_id, rows, sheet_name=None):
        title = sheet_name or self.source_sheet_name
        range_encoded = urllib.parse.quote(f"{title}!A2:AM", safe="")
        resp = self.request_with_auth(
            "POST",
            (
                f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/"
                f"{range_encoded}:append?valueInputOption=RAW&insertDataOption=INSERT_ROWS"
            ),
            json={"majorDimension": "ROWS", "values": rows},
            timeout=60,
            json_body=True,
        )
        return resp.json()

    def export_xlsx(self, spreadsheet_id, local_path):
        local_path = Path(local_path)
        local_path.parent.mkdir(parents=True, exist_ok=True)
        resp = self.request_with_auth(
            "GET",
            f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}/export",
            params={"mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            timeout=120,
        )
        local_path.write_bytes(resp.content)
        return str(local_path)


class CosmetologistHunter:
    def __init__(self, city, firecrawl=None, site_control=None):
        self.city = str(city or "").strip() or "Москва"
        self.city_lc = self.city.lower()
        self.session = requests.Session()
        self.session.headers.update({"user-agent": "Mozilla/5.0"})
        self.firecrawl = firecrawl or FirecrawlClient()
        self.site_control = site_control or SiteControlKitClient()
        self.fetch_trace = deque(maxlen=MAX_FETCH_TRACE_ITEMS)

    def _record_fetch_attempt(self, url, backend, html, required_markers, duration_ms, accepted, reason=""):
        item = {
            "url": url,
            "backend": backend,
            "html_length": len(html or ""),
            "required_markers": [str(marker or "") for marker in (required_markers or []) if str(marker or "")],
            "accepted": bool(accepted),
            "reason": str(reason or "").strip(),
            "duration_ms": int(duration_ms),
            "at": now_text(),
        }
        self.fetch_trace.append(item)
        print(
            "fetch_attempt "
            f"backend={item['backend']} accepted={item['accepted']} reason={item['reason']} "
            f"html_length={item['html_length']} duration_ms={item['duration_ms']} url={item['url']}",
            flush=True,
        )

    def get_fetch_trace(self, limit=30):
        safe_limit = max(1, int(limit or 30))
        items = list(self.fetch_trace)
        return items[-safe_limit:]

    def _preferred_backends(self, url):
        lower_url = str(url or "").lower()
        if "2gis.ru" in lower_url:
            return ["direct", "firecrawl", "site_control"]
        if "prodoctorov.ru" in lower_url:
            if "/vrach/" in lower_url:
                return ["direct"]
            return ["direct", "firecrawl", "site_control"]
        if "yandex." in lower_url:
            return ["direct", "firecrawl", "site_control"]
        return ["direct", "firecrawl", "site_control"]

    def _backend_timeout(self, url, backend):
        lower_url = str(url or "").lower()
        if backend == "direct" and "2gis.ru" in lower_url:
            return 5
        if backend == "direct" and "prodoctorov.ru" in lower_url:
            return 6
        if backend == "direct" and "yandex." in lower_url:
            return 6
        if backend == "firecrawl":
            return 12
        if backend == "site_control":
            return 12
        return 10

    def _matches_required_markers(self, html, required_markers=None):
        if not html:
            return False
        markers = [str(marker or "") for marker in (required_markers or []) if str(marker or "")]
        if not markers:
            return True
        return any(marker in html for marker in markers)

    def _has_blocking_content(self, url, html):
        text = str(html or "")
        if not text:
            return True
        lower_text = text.lower()
        lower_url = str(url or "").lower()
        if "yandex." in lower_url and ("showcaptcha" in lower_text or "form-fb-hint" in lower_text):
            return True
        return False

    def _fetch_via_backend(self, backend, url):
        timeout = self._backend_timeout(url, backend)
        if backend == "firecrawl" and self.firecrawl.enabled():
            return self.firecrawl.scrape_html(url, timeout=timeout)
        if backend == "site_control" and self.site_control.enabled():
            return self.site_control.fetch_html(url, wait_timeout=timeout)
        if backend == "direct":
            try:
                return self.session.get(url, timeout=timeout).text
            except Exception:
                return ""
        return ""

    def fetch_html(self, url, required_markers=None):
        for backend in self._preferred_backends(url):
            started_at = time.time()
            html = self._fetch_via_backend(backend, url)
            duration_ms = int((time.time() - started_at) * 1000)
            if not html:
                self._record_fetch_attempt(
                    url, backend, html, required_markers, duration_ms, accepted=False, reason="empty"
                )
                continue
            if self._has_blocking_content(url, html):
                self._record_fetch_attempt(
                    url, backend, html, required_markers, duration_ms, accepted=False, reason="blocking_content"
                )
                continue
            if self._matches_required_markers(html, required_markers):
                self._record_fetch_attempt(
                    url, backend, html, required_markers, duration_ms, accepted=True, reason="ok"
                )
                return html
            self._record_fetch_attempt(
                url, backend, html, required_markers, duration_ms, accepted=False, reason="missing_markers"
            )
        return ""

    def _candidate_queries(self, max_queries=None):
        out = []
        seen = set()
        limit = max_queries or sum(len(values) for values in SEARCH_STRATEGIES.values())
        ordered_groups = ["private", "doctor", "core"]
        for group in ordered_groups:
            for query in SEARCH_STRATEGIES.get(group, []):
                prepared = f"{query} {self.city}".strip()
                if prepared in seen:
                    continue
                seen.add(prepared)
                out.append(prepared)
                if len(out) >= limit:
                    return out
        return out

    def build_yandex_urls(self, max_queries=None):
        queries = self._candidate_queries(max_queries=max_queries)
        preset = CITY_PRESETS.get(self.city_lc)
        urls = []
        if preset:
            for query in queries:
                for template in preset["yandex_urls"]:
                    urls.append(template.format(query=urllib.parse.quote(query)))
            return urls
        for query in queries:
            urls.append(f"https://yandex.com/maps/?text={urllib.parse.quote(query)}")
        return urls

    def build_2gis_urls(self, max_queries=None, max_pages=3):
        preset = CITY_PRESETS.get(self.city_lc)
        urls = []
        pages = range(1, max(1, int(max_pages or 3)) + 1)
        queries = self._candidate_queries(max_queries=max_queries)
        if preset:
            for query in queries:
                q = urllib.parse.quote(query)
                for page in pages:
                    for template in preset["d2gis_urls"]:
                        urls.append(template.format(query=q, page=page))
            return urls
        for query in queries:
            q = urllib.parse.quote(query)
            for page in pages:
                urls.append(f"https://2gis.ru/search/{q}/page/{page}")
        return urls

    def build_prodoctorov_urls(self, max_pages=3):
        preset = CITY_PRESETS.get(self.city_lc) or {}
        city_slug = str(preset.get("prodoctorov_city") or "").strip()
        if not city_slug:
            return []
        max_pages = max(1, int(max_pages or 3))
        urls = []
        for page in range(1, max_pages + 1):
            suffix = "" if page == 1 else f"?page={page}"
            urls.append(f"https://prodoctorov.ru/{city_slug}/kosmetolog/{suffix}")
        return urls

    def parse_yandex_url(self, url):
        text = self.fetch_html(url, required_markers=['type":"business', '"seoname"', '"phones"'])
        if not text:
            return []
        out = []
        decoder = json.JSONDecoder()
        needle = '{"type":"business"'
        index = 0
        while True:
            idx = text.find(needle, index)
            if idx == -1:
                break
            try:
                obj, end = decoder.raw_decode(text[idx:])
            except Exception:
                index = idx + 1
                continue
            index = idx + max(1, end)
            name = fix_text(obj.get("title") or "") or "Косметолог"
            categories = " ".join(fix_text((item or {}).get("name") or "") for item in (obj.get("categories") or []))
            if not (relevant_text(name) or relevant_text(categories)):
                continue
            phones = unique_phones(
                [
                    (phone_info or {}).get("value") or (phone_info or {}).get("number") or ""
                    for phone_info in (obj.get("phones") or [])
                ]
            )
            if not phones:
                continue
            source_url = url
            item_id = str(obj.get("id") or "").strip()
            seoname = str(obj.get("seoname") or "").strip()
            if item_id and seoname:
                source_url = f"https://yandex.com/maps/org/{seoname}/{item_id}"
            out.append(
                {
                    "company_name": name,
                    "phone_primary": phones[0],
                    "phone_secondary": phones[1] if len(phones) > 1 else "",
                    "address": fix_text(obj.get("fullAddress") or obj.get("address") or ""),
                    "source": "yandex",
                    "source_url": source_url,
                    "city": self.city,
                    "private_hint": categories,
                    "score": score_name(name) + 10,
                }
            )
        return out

    def parse_2gis_search_url(self, url):
        text = self.fetch_html(url, required_markers=["/firm/"])
        if not text:
            return []
        return list(dict.fromkeys(re.findall(r"/firm/(\d+)", text)))

    def parse_2gis_firm(self, firm_id):
        url = f"https://2gis.ru/{'moscow' if self.city_lc == 'москва' else 'search'}/firm/{firm_id}"
        text = self.fetch_html(url, required_markers=["initialState", "contact_groups"])
        if not text:
            return []
        match = re.search(r"var initialState = JSON\.parse\('(.*?)'\);\s+var __REACT_QUERY_STATE__", text, re.S)
        if not match:
            return []
        try:
            payload = json.loads(bytes(match.group(1), "utf-8").decode("unicode_escape"))
        except Exception:
            return []
        stack = [payload]
        profile = None
        while stack:
            current = stack.pop()
            if isinstance(current, dict):
                if "contact_groups" in current and isinstance(current["contact_groups"], list):
                    if str(current.get("id")) == str(firm_id):
                        profile = current
                        break
                    if profile is None:
                        profile = current
                stack.extend(current.values())
            elif isinstance(current, list):
                stack.extend(current)
        if not profile:
            return []
        name = fix_text(profile.get("name") or "") or "Косметолог"
        rubric_aliases = [str((item or {}).get("alias") or "") for item in (profile.get("rubrics") or [])]
        rubric_names = [fix_text((item or {}).get("name") or "") for item in (profile.get("rubrics") or [])]
        relevant = any(alias in STRICT_RUBRIC_ALIASES for alias in rubric_aliases)
        relevant = relevant or relevant_text(name) or any(relevant_text(item) for item in rubric_names)
        if not relevant:
            return []
        phones = unique_phones(
            [
                (contact or {}).get("value") or (contact or {}).get("text") or ""
                for group in (profile.get("contact_groups") or [])
                for contact in ((group or {}).get("contacts") or [])
                if (contact or {}).get("type") == "phone"
            ]
        )
        if not phones:
            return []
        address = fix_text(
            profile.get("full_address_name")
            or profile.get("address_name")
            or profile.get("address")
            or profile.get("address_comment")
            or ""
        )
        return [
            {
                "company_name": name,
                "phone_primary": phones[0],
                "phone_secondary": phones[1] if len(phones) > 1 else "",
                "address": address,
                "source": "2gis",
                "source_url": f"https://2gis.ru/moscow/firm/{firm_id}",
                "city": self.city,
                "private_hint": " ".join(rubric_names),
                "score": score_name(name) + (20 if any(alias == "kosmetolog" for alias in rubric_aliases) else 0),
            }
        ]

    def parse_prodoctorov_list_url(self, url):
        text = self.fetch_html(url, required_markers=["/vrach/", "window.__INITIAL_STATE__"])
        if not text:
            return []
        links = []
        seen = set()
        for match in re.finditer(r'/[a-z0-9_-]+/vrach/\d+-[^"#?]+/', text, re.I):
            raw = match.group(0).replace("%23", "#")
            cleaned = raw.split("#", 1)[0]
            if cleaned in seen:
                continue
            seen.add(cleaned)
            links.append(urllib.parse.urljoin("https://prodoctorov.ru", cleaned))
        return links

    def parse_prodoctorov_profile(self, url):
        text = self.fetch_html(url, required_markers=["window.__INITIAL_STATE__", '"doctor_specialities"'])
        if not text:
            return []
        state = extract_json_object_after(text, "window.__INITIAL_STATE__ = ")
        if not state:
            return []
        doctor = state.get("doctor") or {}
        first_name = fix_text(doctor.get("firstname") or "").strip()
        second_name = fix_text(doctor.get("secondname") or "").strip()
        surname = fix_text(doctor.get("surname") or "").strip()
        full_name = " ".join(part for part in [surname, first_name, second_name] if part).strip() or "Косметолог"
        speciality_names = [
            fix_text((item or {}).get("name") or "")
            for item in (state.get("doctor_specialities") or [])
        ]
        if not any("космет" in item.lower() for item in speciality_names):
            return []

        records = []
        seen = set()
        for item in (state.get("lpu_addresses") or []):
            lpu = (item or {}).get("lpu") or {}
            lpu_name = fix_text(lpu.get("name") or (item or {}).get("name") or "").strip()
            address = fix_text((item or {}).get("address") or lpu.get("address") or "").strip()
            phones = unique_phones(
                [
                    (item or {}).get("phone") or "",
                    ((lpu.get("lpuphone") or {}) if isinstance(lpu.get("lpuphone"), dict) else {}).get("phone") or "",
                ]
            )
            if not phones:
                continue
            company_name = full_name if full_name else (lpu_name or "Косметолог")
            key = (company_name, address, phones[0])
            if key in seen:
                continue
            seen.add(key)
            records.append(
                {
                    "company_name": company_name,
                    "phone_primary": phones[0],
                    "phone_secondary": phones[1] if len(phones) > 1 else "",
                    "address": address,
                    "source": "prodoctorov",
                    "source_url": url,
                    "city": self.city,
                    "private_hint": " ".join(speciality_names),
                    "score": 85 if "космет" in " ".join(speciality_names).lower() else 65,
                }
            )
        return records

    def _candidate_buffer_size(self, count):
        return count if count >= 30 else max(count + 3, count * 2)

    def _collect_candidates(self, count, existing_signatures):
        existing_phones = set(existing_signatures.get("phones") or set())
        existing_names = set(existing_signatures.get("names") or set())
        existing_company_addresses = set(existing_signatures.get("company_addresses") or set())
        buffer_size = self._candidate_buffer_size(count)
        candidates = {}
        seen_names = set()
        seen_company_addresses = set()
        max_queries = 4 if count <= 20 else 8
        max_pages = 1 if count <= 20 else 8
        search_timeout = DEFAULT_SEARCH_TIMEOUT_SECONDS
        if count > 20:
            search_timeout = max(search_timeout, count * 6)
        deadline = time.time() + max(20, search_timeout)

        def timed_out():
            return time.time() >= deadline

        def add_records(items):
            for item in items:
                primary_phone = item.get("phone_primary") or ""
                secondary_phone = item.get("phone_secondary") or ""
                company_key = normalize_company_name(item.get("company_name") or "")
                company_address_key = make_company_address_key(item.get("company_name") or "", item.get("address") or "")
                if not primary_phone:
                    continue
                is_private, private_reason = private_cosmetologist_match(item)
                if not is_private:
                    continue
                item["private_match_reason"] = private_reason
                if primary_phone in existing_phones or primary_phone in candidates:
                    continue
                if secondary_phone and secondary_phone in existing_phones:
                    continue
                if company_key and (company_key in existing_names or company_key in seen_names):
                    continue
                if company_address_key and (
                    company_address_key in existing_company_addresses or company_address_key in seen_company_addresses
                ):
                    continue
                candidates[primary_phone] = item
                existing_phones.add(primary_phone)
                if secondary_phone:
                    existing_phones.add(secondary_phone)
                if company_key:
                    seen_names.add(company_key)
                if company_address_key:
                    seen_company_addresses.add(company_address_key)

        if not timed_out():
            profile_links = []
            profile_seen = set()
            max_profile_links = max(buffer_size * 3, 60)
            for url in self.build_prodoctorov_urls(max_pages=max_pages):
                if timed_out():
                    break
                for profile_url in self.parse_prodoctorov_list_url(url):
                    if profile_url in profile_seen:
                        continue
                    profile_seen.add(profile_url)
                    profile_links.append(profile_url)
                    if len(profile_links) >= max_profile_links:
                        break
                if len(profile_links) >= max_profile_links:
                    break

            threads = []
            lock = threading.Lock()

            def profile_worker(profile_url):
                if timed_out():
                    return
                try:
                    records = self.parse_prodoctorov_profile(profile_url)
                    if records:
                        with lock:
                            add_records(records)
                except Exception:
                    return

            for profile_url in profile_links:
                if timed_out():
                    break
                thread = threading.Thread(target=profile_worker, args=(profile_url,))
                thread.start()
                threads.append(thread)
                if len(threads) >= 8:
                    for active in threads:
                        active.join(timeout=8)
                    threads = []
                    if len(candidates) >= buffer_size:
                        return list(candidates.values())
            for active in threads:
                active.join(timeout=8)

        print(f"candidate_stage stage=prodoctorov count={len(candidates)}", flush=True)
        if len(candidates) >= buffer_size or timed_out():
            return list(candidates.values())

        for url in self.build_yandex_urls(max_queries=max_queries):
            if timed_out():
                break
            add_records(self.parse_yandex_url(url))
            if len(candidates) >= buffer_size:
                return list(candidates.values())

        print(f"candidate_stage stage=yandex count={len(candidates)}", flush=True)
        if len(candidates) >= buffer_size or timed_out():
            return list(candidates.values())
        if count >= 30:
            return list(candidates.values())

        firm_ids = []
        firm_seen = set()
        max_firm_ids = max(buffer_size * 2, 16)
        for url in self.build_2gis_urls(max_queries=max_queries, max_pages=max_pages):
            if timed_out():
                break
            for firm_id in self.parse_2gis_search_url(url):
                if firm_id in firm_seen:
                    continue
                firm_seen.add(firm_id)
                firm_ids.append(firm_id)
                if len(firm_ids) >= max_firm_ids:
                    break
            if len(firm_ids) >= max_firm_ids:
                break

        threads = []
        lock = threading.Lock()

        def worker(fid):
            if timed_out():
                return
            try:
                records = self.parse_2gis_firm(fid)
                if records:
                    with lock:
                        add_records(records)
            except Exception:
                return

        for firm_id in firm_ids:
            if timed_out():
                break
            thread = threading.Thread(target=worker, args=(firm_id,))
            thread.start()
            threads.append(thread)
            if len(threads) >= 4:
                for active in threads:
                    active.join(timeout=14)
                threads = []
                if len(candidates) >= buffer_size:
                    return list(candidates.values())
        for active in threads:
            active.join(timeout=14)

        return list(candidates.values())

    def find_contacts(self, count, existing_signatures):
        records = self._collect_candidates(count, existing_signatures)
        records.sort(key=lambda item: (-int(item.get("score", 0)), item.get("company_name", ""), item["phone_primary"]))
        return records[:count]


class HunterService:
    def __init__(
        self,
        source_spreadsheet_id=DEFAULT_SOURCE_SPREADSHEET_ID,
        source_sheet_name=DEFAULT_SHEET_NAME,
        local_output_dir=DEFAULT_LOCAL_OUTPUT_DIR,
        template_xlsx=DEFAULT_TEMPLATE_XLSX,
        settings_path=DEFAULT_SETTINGS_PATH,
        preview_dir=DEFAULT_PREVIEW_DIR,
        drive_folder_id=DEFAULT_DRIVE_FOLDER_ID,
    ):
        self.google = GoogleSheetsClient(source_spreadsheet_id, source_sheet_name, drive_folder_id=drive_folder_id)
        self.firecrawl = FirecrawlClient(
            base_url=DEFAULT_FIRECRAWL_BASE_URL,
            api_key=DEFAULT_FIRECRAWL_API_KEY,
            root_path=DEFAULT_FIRECRAWL_ROOT,
        )
        self.site_control = SiteControlKitClient(
            server_url=DEFAULT_SITE_CONTROL_SERVER_URL,
            token=DEFAULT_SITE_CONTROL_TOKEN,
            client_id=DEFAULT_SITE_CONTROL_CLIENT_ID,
            root_path=DEFAULT_SITE_CONTROL_ROOT,
            browser_service=DEFAULT_SITE_CONTROL_BROWSER_SERVICE,
            hub_service=DEFAULT_SITE_CONTROL_HUB_SERVICE,
            browser_timer=DEFAULT_SITE_CONTROL_BROWSER_TIMER,
            on_demand=DEFAULT_SITE_CONTROL_ON_DEMAND,
        )
        self.source_spreadsheet_id = source_spreadsheet_id
        self.source_sheet_name = source_sheet_name
        self.local_output_dir = Path(local_output_dir)
        self.template_xlsx = Path(template_xlsx)
        self.settings_path = settings_path
        self.preview_dir = Path(preview_dir)
        self.drive_folder_id = str(drive_folder_id or "").strip()
        self.lock = threading.Lock()
        self.last_fetch_trace = []
        self._debug_summary_cache = {}

    def tooling_status(self):
        return {
            "server_tool_root": DEFAULT_SERVER_TOOL_ROOT,
            "firecrawl": self.firecrawl.status(),
            "site_control": self.site_control.status(),
        }

    def fetch_trace_status(self, limit=30):
        safe_limit = max(1, min(int(limit or 30), MAX_FETCH_TRACE_ITEMS))
        items = self.last_fetch_trace[-safe_limit:]
        return {
            "items": items,
            "count": len(items),
        }

    def debug_summary(self, chat_id="", city="", estimate_cap=30, refresh=False):
        settings = self.get_settings(chat_id)
        resolved_city = str(city or settings.get("city") or "Москва").strip()
        estimate_cap = max(1, min(int(estimate_cap or 30), 120))
        cache_key = (resolved_city.lower(), estimate_cap)
        cached = self._debug_summary_cache.get(cache_key) or {}
        if not refresh and cached and (time.time() - float(cached.get("at_ts") or 0)) < 120:
            return cached.get("payload") or {}
        existing_signatures = self.google.existing_signatures_for_city(resolved_city)
        hunter = CosmetologistHunter(resolved_city, firecrawl=self.firecrawl, site_control=self.site_control)
        try:
            contacts = hunter.find_contacts(estimate_cap, existing_signatures)
        finally:
            hunter.site_control.shutdown()
        payload = {
            "city": resolved_city,
            "estimate_cap": estimate_cap,
            "estimated_available": len(contacts),
            "sample": contacts[:3],
        }
        self._debug_summary_cache[cache_key] = {
            "at_ts": time.time(),
            "payload": payload,
        }
        return payload

    def load_settings(self):
        return load_json_file(self.settings_path, {})

    def save_settings(self, payload):
        save_json_file(self.settings_path, payload)

    def get_settings(self, chat_id):
        payload = self.load_settings()
        chat_key = str(chat_id or "").strip()
        default = {
            "city": "Москва",
            "count": 45,
            "source_spreadsheet_id": self.source_spreadsheet_id,
        }
        if not chat_key:
            return default
        return {**default, **payload.get(chat_key, {})}

    def set_settings(self, chat_id, city=None, count=None, source_spreadsheet_id=None):
        payload = self.load_settings()
        chat_key = str(chat_id or "").strip()
        if not chat_key:
            raise ValueError("chat_id is required")
        current = payload.get(chat_key, {})
        if city:
            current["city"] = str(city).strip()
        if count is not None:
            current["count"] = int(count)
        if source_spreadsheet_id:
            current["source_spreadsheet_id"] = str(source_spreadsheet_id).strip()
        payload[chat_key] = current
        self.save_settings(payload)
        return self.get_settings(chat_key)

    def write_local_template_copy(self, title, rows):
        target = self.local_output_dir / f"{title}.xlsx"
        self.local_output_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.template_xlsx, target)
        workbook = load_workbook(target)
        worksheet = workbook[self.source_sheet_name]
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = None
        for row_index, row_values in enumerate(rows, start=2):
            for col_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=col_index).value = value
        workbook.save(target)
        return str(target)

    def run_hunt(self, city=None, count=None, chat_id=None, dry_run=False):
        settings = self.get_settings(chat_id)
        city = str(city or settings.get("city") or "Москва").strip()
        count = int(count or settings.get("count") or 45)
        if count <= 0:
            raise ValueError("count must be positive")

        with self.lock:
            title_prefix = build_title_prefix(city)
            ordinal = self.google.next_ordinal(title_prefix)
            title = build_sheet_title(city, ordinal)
            existing_signatures = self.google.existing_signatures_for_city(city)
            hunter = CosmetologistHunter(city, firecrawl=self.firecrawl, site_control=self.site_control)
            try:
                contacts = hunter.find_contacts(count, existing_signatures)
                self.last_fetch_trace = hunter.get_fetch_trace(limit=MAX_FETCH_TRACE_ITEMS)
            finally:
                hunter.site_control.shutdown()
            if len(contacts) < count:
                raise RuntimeError(
                    f"Недостаточно новых контактов косметологов для города {city}: найдено {len(contacts)}, нужно {count}"
                )
            rows = build_output_rows(contacts)
            preview_path = self.preview_dir / f"{title}.json"
            save_json_file(preview_path, contacts)

            if dry_run:
                return {
                    "ok": True,
                    "dry_run": True,
                    "city": city,
                    "count": count,
                    "title": title,
                    "ordinal": ordinal,
                    "preview_path": str(preview_path),
                    "contacts": contacts,
                    "fetch_trace": self.last_fetch_trace[-20:],
                }

            spreadsheet_id = self.google.copy_sheet(title)
            self.google.clear_data_rows(spreadsheet_id, self.source_sheet_name)
            self.google.append_rows(spreadsheet_id, rows, self.source_sheet_name)
            local_file = self.local_output_dir / f"{title}.xlsx"
            try:
                self.google.export_xlsx(spreadsheet_id, local_file)
            except Exception:
                self.write_local_template_copy(title, rows)
            return {
                "ok": True,
                "city": city,
                "count": count,
                "title": title,
                "ordinal": ordinal,
                "spreadsheet_id": spreadsheet_id,
                "google_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
                "local_file": str(local_file),
                "preview_path": str(preview_path),
                "contacts_count": len(contacts),
            }


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


class HunterRequestHandler(BaseHTTPRequestHandler):
    service = None

    def _authorized(self):
        expected = os.getenv("COSMETOLOGIST_HUNTER_AUTH_TOKEN", "").strip()
        if not expected:
            return True
        received = self.headers.get("Authorization", "").strip()
        return received == f"Bearer {expected}"

    def _send(self, code, payload):
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        try:
            self.wfile.write(raw)
        except BrokenPipeError:
            return

    def _read_json(self):
        length = int(self.headers.get("Content-Length", "0") or 0)
        if length <= 0:
            return {}
        body = self.rfile.read(length).decode("utf-8")
        return json.loads(body) if body else {}

    def do_GET(self):
        try:
            if not self._authorized():
                self._send(401, {"ok": False, "error": "Unauthorized"})
                return
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == "/health":
                self._send(200, {"ok": True, "service": "cosmetologist_hunter"})
                return
            if parsed.path == "/settings/get":
                params = urllib.parse.parse_qs(parsed.query)
                chat_id = (params.get("chat_id") or [""])[0]
                self._send(200, {"ok": True, "settings": self.service.get_settings(chat_id)})
                return
            if parsed.path == "/tooling/status":
                self._send(200, {"ok": True, "tooling": self.service.tooling_status()})
                return
            if parsed.path == "/debug/fetch-trace":
                params = urllib.parse.parse_qs(parsed.query)
                limit = (params.get("limit") or ["30"])[0]
                self._send(200, {"ok": True, "fetch_trace": self.service.fetch_trace_status(limit=limit)})
                return
            if parsed.path == "/debug/summary":
                params = urllib.parse.parse_qs(parsed.query)
                chat_id = (params.get("chat_id") or [""])[0]
                city = (params.get("city") or [""])[0]
                estimate_cap = (params.get("estimate_cap") or ["30"])[0]
                refresh = ((params.get("refresh") or ["0"])[0]).strip().lower() in {"1", "true", "yes", "on"}
                self._send(
                    200,
                    {
                        "ok": True,
                        "summary": self.service.debug_summary(
                            chat_id=chat_id,
                            city=city,
                            estimate_cap=estimate_cap,
                            refresh=refresh,
                        ),
                    },
                )
                return
            self._send(404, {"ok": False, "error": "Not found"})
        except Exception as exc:
            self._send(500, {"ok": False, "error": str(exc)})

    def do_POST(self):
        try:
            if not self._authorized():
                self._send(401, {"ok": False, "error": "Unauthorized"})
                return
            parsed = urllib.parse.urlparse(self.path)
            payload = self._read_json()
            if parsed.path == "/settings/set":
                settings = self.service.set_settings(
                    chat_id=payload.get("chat_id"),
                    city=payload.get("city"),
                    count=payload.get("count"),
                    source_spreadsheet_id=payload.get("source_spreadsheet_id"),
                )
                self._send(200, {"ok": True, "settings": settings})
                return
            if parsed.path == "/run":
                result = self.service.run_hunt(
                    city=payload.get("city"),
                    count=payload.get("count"),
                    chat_id=payload.get("chat_id"),
                    dry_run=bool(payload.get("dry_run")),
                )
                self._send(200, result)
                return
            self._send(404, {"ok": False, "error": "Not found"})
        except Exception as exc:
            self._send(500, {"ok": False, "error": str(exc)})

    def log_message(self, format, *args):
        return


def run_server(args):
    service = HunterService(
        source_spreadsheet_id=args.source_spreadsheet_id,
        source_sheet_name=args.sheet_name,
        local_output_dir=args.local_output_dir,
        template_xlsx=args.template_xlsx,
        settings_path=args.settings_path,
        preview_dir=args.preview_dir,
        drive_folder_id=args.drive_folder_id,
    )
    HunterRequestHandler.service = service
    server = ThreadingHTTPServer((args.host, args.port), HunterRequestHandler)
    print(f"Cosmetologist hunter service listening on http://{args.host}:{args.port}")
    server.serve_forever()


def run_once(args):
    service = HunterService(
        source_spreadsheet_id=args.source_spreadsheet_id,
        source_sheet_name=args.sheet_name,
        local_output_dir=args.local_output_dir,
        template_xlsx=args.template_xlsx,
        settings_path=args.settings_path,
        preview_dir=args.preview_dir,
        drive_folder_id=args.drive_folder_id,
    )
    result = service.run_hunt(city=args.city, count=args.count, chat_id=args.chat_id, dry_run=args.dry_run)
    print(json.dumps(result, ensure_ascii=False, indent=2))


def build_parser():
    parser = argparse.ArgumentParser(description="Cosmetologist hunter agent service")
    parser.add_argument("--source-spreadsheet-id", default=DEFAULT_SOURCE_SPREADSHEET_ID)
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--local-output-dir", default=DEFAULT_LOCAL_OUTPUT_DIR)
    parser.add_argument("--template-xlsx", default=DEFAULT_TEMPLATE_XLSX)
    parser.add_argument("--settings-path", default=DEFAULT_SETTINGS_PATH)
    parser.add_argument("--preview-dir", default=DEFAULT_PREVIEW_DIR)
    parser.add_argument("--drive-folder-id", default=DEFAULT_DRIVE_FOLDER_ID)

    subparsers = parser.add_subparsers(dest="command", required=True)

    serve = subparsers.add_parser("serve", help="Run HTTP service")
    serve.add_argument("--host", default="127.0.0.1")
    serve.add_argument("--port", type=int, default=DEFAULT_PORT)
    serve.set_defaults(handler=run_server)

    run = subparsers.add_parser("run", help="Run one search job")
    run.add_argument("--city", default="Москва")
    run.add_argument("--count", type=int, default=45)
    run.add_argument("--chat-id", default="")
    run.add_argument("--dry-run", action="store_true")
    run.set_defaults(handler=run_once)
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    args.handler(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
